import json
import os
import re
import sys
import time
import openpyxl
import requests
import threading
from loguru import logger
from retry import retry


def norm_str(str):
    new_str = re.sub(r"|[\\/:*?\"<>| ]+", "", str).replace('\n', '').replace('\r', '')
    return new_str

def norm_text(text):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
    return text


def timestamp_to_str(timestamp):
    time_local = time.localtime(timestamp / 1000)
    dt = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
    return dt

def handle_user_info(data, user_id):
    home_url = f'https://www.xiaohongshu.com/user/profile/{user_id}'
    nickname = data['basic_info']['nickname']
    avatar = data['basic_info']['imageb']
    red_id = data['basic_info']['red_id']
    gender = data['basic_info']['gender']
    if gender == 0:
        gender = '男'
    elif gender == 1:
        gender = '女'
    else:
        gender = '未知'
    ip_location = data['basic_info']['ip_location']
    desc = data['basic_info']['desc']
    follows = data['interactions'][0]['count']
    fans = data['interactions'][1]['count']
    interaction = data['interactions'][2]['count']
    tags_temp = data['tags']
    tags = []
    for tag in tags_temp:
        try:
            tags.append(tag['name'])
        except:
            pass
    return {
        'user_id': user_id,
        'home_url': home_url,
        'nickname': nickname,
        'avatar': avatar,
        'red_id': red_id,
        'gender': gender,
        'ip_location': ip_location,
        'desc': desc,
        'follows': follows,
        'fans': fans,
        'interaction': interaction,
        'tags': tags,
    }

def handle_note_info(data):
    note_id = data['id']
    note_url = data['url']
    note_type = data['note_card']['type']
    if note_type == 'normal':
        note_type = '图集'
    else:
        note_type = '视频'
    user_id = data['note_card']['user']['user_id']
    home_url = f'https://www.xiaohongshu.com/user/profile/{user_id}'
    nickname = data['note_card']['user']['nickname']
    avatar = data['note_card']['user']['avatar']
    title = data['note_card']['title']
    if title.strip() == '':
        title = f'无标题'
    desc = data['note_card']['desc']
    liked_count = data['note_card']['interact_info']['liked_count']
    collected_count = data['note_card']['interact_info']['collected_count']
    comment_count = data['note_card']['interact_info']['comment_count']
    share_count = data['note_card']['interact_info']['share_count']
    image_list_temp = data['note_card']['image_list']
    image_list = []
    for image in image_list_temp:
        try:
            image_list.append(image['info_list'][1]['url'])
            # success, msg, img_url = XHS_Apis.get_note_no_water_img(image['info_list'][1]['url'])
            # image_list.append(img_url)
        except:
            pass
    if note_type == '视频':
        try:
            video_cover = image_list[0] if image_list else None
            # Safe access to video key
            video_key = data['note_card']['video']['consumer']['origin_video_key']
            video_addr = 'https://sns-video-bd.xhscdn.com/' + video_key
        except Exception as e:
            logger.warning(f"Failed to extract video address: {e}")
            video_cover = None
            video_addr = None
        # success, msg, video_addr = XHS_Apis.get_note_no_water_video(note_id)
    else:
        video_cover = None
        video_addr = None
    tags_temp = data['note_card']['tag_list']
    tags = []
    for tag in tags_temp:
        try:
            tags.append(tag['name'])
        except:
            pass
    upload_time = timestamp_to_str(data['note_card']['time'])
    if 'ip_location' in data['note_card']:
        ip_location = data['note_card']['ip_location']
    else:
        ip_location = '未知'
    return {
        'note_id': note_id,
        'note_url': note_url,
        'note_type': note_type,
        'user_id': user_id,
        'home_url': home_url,
        'nickname': nickname,
        'avatar': avatar,
        'title': title,
        'desc': desc,
        'liked_count': liked_count,
        'collected_count': collected_count,
        'comment_count': comment_count,
        'share_count': share_count,
        'video_cover': video_cover,
        'video_addr': video_addr,
        'image_list': image_list,
        'tags': tags,
        'upload_time': upload_time,
        'ip_location': ip_location,
    }

def handle_comment_info(data):
    note_id = data['note_id']
    note_url = data['note_url']
    comment_id = data['id']
    user_id = data['user_info']['user_id']
    home_url = f'https://www.xiaohongshu.com/user/profile/{user_id}'
    nickname = data['user_info']['nickname']
    avatar = data['user_info']['image']
    content = data['content']
    show_tags = data['show_tags']
    like_count = data['like_count']
    upload_time = timestamp_to_str(data['create_time'])
    try:
        ip_location = data['ip_location']
    except:
        ip_location = '未知'
    pictures = []
    try:
        pictures_temp = data['pictures']
        for picture in pictures_temp:
            try:
                pictures.append(picture['info_list'][1]['url'])
                # success, msg, img_url = XHS_Apis.get_note_no_water_img(picture['info_list'][1]['url'])
                # pictures.append(img_url)
            except:
                pass
    except:
        pass
    return {
        'note_id': note_id,
        'note_url': note_url,
        'comment_id': comment_id,
        'user_id': user_id,
        'home_url': home_url,
        'nickname': nickname,
        'avatar': avatar,
        'content': content,
        'show_tags': show_tags,
        'like_count': like_count,
        'upload_time': upload_time,
        'ip_location': ip_location,
        'pictures': pictures,
    }
def save_to_xlsx(datas, file_path, type='note'):
    """Save data to xlsx file, appending to existing file or creating new one.
    Handles corrupted xlsx files gracefully by creating a new workbook."""
    def _create_new_workbook(type):
        wb = openpyxl.Workbook()
        ws = wb.active
        if type == 'note':
            headers = ['笔记id', '笔记url', '笔记类型', '用户id', '用户主页url', '昵称', '头像url', '标题', '描述', '点赞数量', '收藏数量', '评论数量', '分享数量', '视频封面url', '视频地址url', '图片地址url列表', '标签', '上传时间', 'ip归属地']
        elif type == 'user':
            headers = ['用户id', '用户主页url', '用户名', '头像url', '小红书号', '性别', 'ip地址', '介绍', '关注数量', '粉丝数量', '作品被赞和收藏数量', '标签']
        else:
            headers = ['笔记id', '笔记url', '评论id', '用户id', '用户主页url', '昵称', '头像url', '评论内容', '评论标签', '点赞数量', '上传时间', 'ip归属地', '图片地址url列表']
        ws.append(headers)
        return wb, ws

    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            # Excel file is corrupted; back it up and create a fresh one
            logger.warning(f"Corrupted xlsx file {file_path}: {e}. Backing up and creating new.")
            backup_path = file_path + '.corrupted'
            try:
                os.rename(file_path, backup_path)
            except Exception:
                pass
            wb, ws = _create_new_workbook(type)
    else:
        wb, ws = _create_new_workbook(type)
    
    for data in datas:
        data = {k: norm_text(str(v)) for k, v in data.items()}
        ws.append(list(data.values()))
    wb.save(file_path)
    logger.info(f'数据保存(追加)至 {file_path}')

def get_scraped_note_ids(file_path):
    if not os.path.exists(file_path):
        return set()
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        ids = set()
        # Iterate from row 2, column 1 (assuming it's '笔记id')
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                ids.add(str(row[0]))
        return ids
    except Exception as e:
        logger.error(f"Error reading existing excel {file_path}: {e}")
        return set()


def save_user_detail(user, path):
    with open(f'{path}/detail.txt', mode="w", encoding="utf-8") as f:
        # 逐行输出到txt里
        f.write(f"用户id: {user['user_id']}\n")
        f.write(f"用户主页url: {user['home_url']}\n")
        f.write(f"用户名: {user['nickname']}\n")
        f.write(f"头像url: {user['avatar']}\n")
        f.write(f"小红书号: {user['red_id']}\n")
        f.write(f"性别: {user['gender']}\n")
        f.write(f"ip地址: {user['ip_location']}\n")
        f.write(f"介绍: {user['desc']}\n")
        f.write(f"关注数量: {user['follows']}\n")
        f.write(f"粉丝数量: {user['fans']}\n")
        f.write(f"作品被赞和收藏数量: {user['interaction']}\n")
        f.write(f"标签: {user['tags']}\n")

def save_note_detail(note, path):
    with open(f'{path}/detail.txt', mode="w", encoding="utf-8") as f:
        # 逐行输出到txt里
        f.write(f"笔记id: {note['note_id']}\n")
        f.write(f"笔记url: {note['note_url']}\n")
        f.write(f"笔记类型: {note['note_type']}\n")
        f.write(f"用户id: {note['user_id']}\n")
        f.write(f"用户主页url: {note['home_url']}\n")
        f.write(f"昵称: {note['nickname']}\n")
        f.write(f"头像url: {note['avatar']}\n")
        f.write(f"标题: {note['title']}\n")
        f.write(f"描述: {note['desc']}\n")
        f.write(f"点赞数量: {note['liked_count']}\n")
        f.write(f"收藏数量: {note['collected_count']}\n")
        f.write(f"评论数量: {note['comment_count']}\n")
        f.write(f"分享数量: {note['share_count']}\n")
        f.write(f"视频封面url: {note['video_cover']}\n")
        f.write(f"视频地址url: {note['video_addr']}\n")
        f.write(f"图片地址url列表: {note['image_list']}\n")
        f.write(f"标签: {note['tags']}\n")
        f.write(f"上传时间: {note['upload_time']}\n")
        f.write(f"ip归属地: {note['ip_location']}\n")
        f.write(f"图片文字: {note.get('ocr_content', '无')}\n")



# Initialize OCR (Global to avoid reloading)
OCR_ENGINE = None
try:
    from paddleocr import PaddleOCR
    # Suppress heavy logging
    import logging
    logging.getLogger("ppocr").setLevel(logging.ERROR)
    OCR_ENGINE = PaddleOCR(use_angle_cls=True, lang="ch")
    logger.info("PaddleOCR initialized successfully.")
except ImportError:
    logger.warning("PaddleOCR not found. OCR features disabled.")
except Exception as e:
    logger.warning(f"Failed to init OCR: {e}")

def is_text_image(img_path, threshold=75.0):
    """Determine if an image is a text-on-solid-background image suitable for OCR.
    
    Analyzes the dominant color percentage of the image. Text post images from
    Xiaohongshu typically have a solid color background (>75% of pixels are one
    color), while photos/screenshots/charts have much more color diversity.
    
    Args:
        img_path: Path to the image file.
        threshold: Minimum dominant color percentage to classify as text image.
                   Default 75% avoids borderline cases that cause slow OCR.
    Returns:
        True if image appears to be a text-on-solid-background image.
    """
    try:
        from PIL import Image
        import numpy as np
        img = Image.open(img_path).convert("RGB").resize((100, 100))
        pixels = np.array(img).reshape(-1, 3)
        # Quantize colors to 16-level bins to group similar shades
        quantized = (pixels // 16) * 16
        _, counts = np.unique(quantized, axis=0, return_counts=True)
        dominant_pct = counts.max() / counts.sum() * 100
        return dominant_pct >= threshold
    except Exception as e:
        logger.warning(f"is_text_image check failed for {img_path}: {e}")
        return True  # Default to True (run OCR) on error


def run_ocr_subprocess(img_path, timeout=60):
    """Run OCR on an image in a separate subprocess with timeout.
    
    Uses Popen + explicit SIGKILL to ensure the subprocess is truly killed on
    timeout, even if it's in uninterruptible sleep (UE+ state). This is more
    robust than subprocess.run(timeout=) which may fail to kill stuck processes.
    
    Args:
        img_path: Path to the image file.
        timeout: Maximum seconds to wait for OCR. Default 60s.
    Returns:
        List of extracted text strings, or None on failure/timeout.
    """
    import subprocess as sp
    import signal
    script_path = os.path.join(os.path.dirname(__file__), 'ocr_subprocess.py')
    python_path = sys.executable
    proc = None
    try:
        proc = sp.Popen(
            [python_path, script_path, img_path],
            stdout=sp.PIPE, stderr=sp.PIPE, text=True,
            preexec_fn=os.setsid  # Create new process group for reliable kill
        )
        stdout, stderr = proc.communicate(timeout=timeout)
        if proc.returncode == 0 and stdout.strip():
            data = json.loads(stdout.strip())
            if 'texts' in data:
                return data['texts']
            else:
                logger.warning(f"OCR subprocess error: {data.get('error', 'unknown')}")
                return None
        else:
            logger.warning(f"OCR subprocess failed (rc={proc.returncode}): {stderr[:200] if stderr else 'no stderr'}")
            return None
    except sp.TimeoutExpired:
        logger.warning(f"OCR subprocess timed out after {timeout}s for {img_path}")
        if proc:
            try:
                # Kill entire process group with SIGKILL
                os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
            except (ProcessLookupError, OSError):
                pass
            proc.wait()  # Reap zombie
        return None
    except Exception as e:
        logger.warning(f"OCR subprocess exception: {e}")
        if proc and proc.poll() is None:
            try:
                os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
            except (ProcessLookupError, OSError):
                pass
        return None

@retry(tries=3, delay=1)
def download_note(note_info, path, save_choice):
    """Download a note's media and perform OCR on text images."""
    note_id = note_info['note_id']
    user_id = note_info['user_id']
    title = note_info['title']
    title = norm_str(title)[:40]
    nickname = note_info['nickname']
    nickname = norm_str(nickname)[:20]
    if title.strip() == '':
        title = f'无标题'
    save_path = f'{path}/{nickname}_{user_id}/{title}_{note_id}'
    check_and_create_path(save_path)
    
    note_type = note_info['note_type']
    ocr_results = []

    # 1. Save Initial Metadata (Ensures persistence even if next steps hang)
    if 'ocr_content' not in note_info:
        note_info['ocr_content'] = "无"
    
    try:
        with open(f'{save_path}/info.json', mode='w', encoding='utf-8') as f:
            f.write(json.dumps(note_info, ensure_ascii=False) + '\n')
        save_note_detail(note_info, save_path)
    except Exception as e:
        logger.error(f"Failed to save initial detail for {note_id}: {e}")

    # 2. Download Media
    if note_type == '图集' and save_choice in ['media', 'media-image', 'all']:
        for img_index, img_url in enumerate(note_info['image_list']):
            img_name = f'image_{img_index}'
            download_media(save_path, img_name, img_url, 'image')
            
            # OCR Logic via subprocess (can be truly killed on timeout)
            if OCR_ENGINE:
                try:
                    img_full_path = os.path.join(save_path, f'{img_name}.jpg')
                    if os.path.exists(img_full_path):
                        # Skip non-text images (photos, screenshots, charts)
                        if not is_text_image(img_full_path):
                            logger.info(f"Skipping OCR for {img_name}: not a text image")
                            ocr_results.append(f"图{img_index+1}: 非文字图(跳过)")
                            continue
                        # Run OCR in subprocess with 60s timeout
                        text_lines = run_ocr_subprocess(img_full_path, timeout=60)
                        if text_lines is not None:
                            content = ", ".join(text_lines) if text_lines else "无"
                        else:
                            content = "OCR超时/失败"
                        
                        ocr_results.append(f"图{img_index+1}: {content}")
                except Exception as e:
                    logger.warning(f"OCR wrapper failed for {img_name}: {e}")

    elif note_type == '视频' and save_choice in ['media', 'media-video', 'all']:
        try:
            if note_info.get('video_cover'):
                download_media(save_path, 'cover', note_info['video_cover'], 'image')
            if note_info.get('video_addr'):
                download_media(save_path, 'video', note_info['video_addr'], 'video')
            else:
                logger.warning(f"Video address is None for {note_id}, skipping video download.")
        except Exception as e:
            logger.error(f"Failed to download video media: {e}")

    # 3. Update Metadata with OCR Content
    if ocr_results:
        note_info['ocr_content'] = " | ".join(ocr_results)
        try:
            with open(f'{save_path}/info.json', mode='w', encoding='utf-8') as f:
                f.write(json.dumps(note_info, ensure_ascii=False) + '\n')
            save_note_detail(note_info, save_path)
        except Exception as e:
            logger.error(f"Failed to update detail for {note_id}: {e}")
    
    return save_path

def download_media(path, name, url, type):
    if not url:
        logger.warning(f"Skipping download for {name}, URL is empty.")
        return
    try:
        if type == 'image':
            content = requests.get(url, timeout=10).content
            with open(path + '/' + name + '.jpg', mode="wb") as f:
                f.write(content)
        elif type == 'video':
            res = requests.get(url, stream=True, timeout=20)
            size = 0
            chunk_size = 1024 * 1024
            with open(path + '/' + name + '.mp4', mode="wb") as f:
                for data in res.iter_content(chunk_size=chunk_size):
                    f.write(data)
                    size += len(data)
    except Exception as e:
        logger.warning(f"Media download failed for {name}: {e}")


def check_and_create_path(path):
    if not os.path.exists(path):
        os.makedirs(path)
