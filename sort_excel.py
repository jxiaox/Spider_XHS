# -*- coding: utf-8 -*-
import openpyxl
import os

file_path = '/Users/jxiaox/github.com/Spider_XHS/datas/excel_datas/5b6150c56b58b741e26b8c7f.xlsx'

def sort_excel():
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Get all rows as a list of values
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        print("Empty file")
        return

    header = rows[0]
    data = rows[1:]

    # Find index of '上传时间'
    try:
        date_idx = header.index('上传时间')
        print(f"Found '上传时间' at column index {date_idx}")
    except ValueError:
        print("Column '上传时间' not found")
        return

    # Sort data. Newest first -> Descending.
    # Data is likely string 'YYYY-MM-DD HH:MM:SS', which sorts correctly.
    print(f"Sorting {len(data)} rows...")
    data.sort(key=lambda x: str(x[date_idx]) if x[date_idx] else '', reverse=True)

    # Create new workbook to be clean/safe
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(header)
    
    for row in data:
        new_ws.append(row)
        
    new_wb.save(file_path)
    print(f"Sorted and saved to {file_path}")

if __name__ == '__main__':
    sort_excel()
